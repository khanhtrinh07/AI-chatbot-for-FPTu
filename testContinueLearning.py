import os
from openpyxl import Workbook, load_workbook
import pandas as pd


def saveNoRe(inp):
    #inp = input("You: ")

    noRe = "D:\Code\Project\\noRp\\noResponses.xlsx"

    if not os.path.exists(noRe):
        # Tạo file mới
        workbook = Workbook()
        worksheet = workbook.active
        # Đặt tên cho các cột
        worksheet.cell(row=1, column=1).value = "Tag"
        worksheet.cell(row=1, column=2).value = "Patterns"
        worksheet.cell(row=1, column=3).value = "Responses"
        workbook.save(noRe)
    else:
        # Mở file có sẵn
        workbook = load_workbook(filename=noRe)
        worksheet = workbook.active

    row_num = worksheet.max_row+1
    # Thêm câu mới vào file
    worksheet.cell(row=row_num, column=2).value = inp

    row_num += 1
    # Lưu file workbook
    workbook.save(noRe)
def collectData():
    mess = input("Bạn có thời gian để hỗ trợ chúng tôi chứ?(Y/N)\nY: Đồng ý\nN: Không đồng ý\nLựa chọn của bạn: ")
    if (mess.lower()=="y"):
        df = pd.read_excel('D:\Code\Project\\noRp\\noResponses.xlsx')
        loop = True
        while loop:
            b_rows = df[df.iloc[:, 1].notnull()].index.tolist()
            if len(b_rows) == 0:
                loop = False
                break
            for index in b_rows:
                # Kiểm tra nếu giá trị ở cột C và cột A đều là null
                if pd.isnull(df.iloc[index, 2]) and pd.isnull(df.iloc[index, 0]):
                    # In giá trị của cột B
                    print("Đây là câu hỏi chưa có câu trả lời:", df.iloc[index, 1])  
                    # Nhập giá trị của cột A và cột C tương ứng với cột B đó
                    a = input('BẠN CÓ THỂ NHẤN \'continue\' ĐỂ BỎ QUA VÀ \'quit\' ĐỂ THOÁT\nBạn muốn gán nhãn cho câu này ở lĩnh vực nào?(Cần sự ngắn gọn và chính xác): ')
                    df.iloc[index, 0] = a
                    if(a.lower()=='continue'):
                        continue
                    if(a.lower()=='quit'):
                        quit()
                    df.to_excel('D:\Code\Project\\noRp\\noResponses.xlsx', index=False)
                    c = input('Bạn có đề xuất câu trả lời cho câu hỏi này?: ')
                    df.iloc[index, 2] = c
                    if(c.lower()=='continue'):
                        continue
                    if(c.lower()=='quit'):
                        quit()
                    df.to_excel('D:\Code\Project\\noRp\\noResponses.xlsx', index=False)
                    answer = input('Bạn có muốn nhập giá trị cho cột B tiếp theo không? (y/n): ')
                    if answer.lower() == 'n':
                        loop = False
                        break
        
        df.to_excel('D:\Code\Project\\noRp\\noResponses.xlsx', index=False)        
    elif(mess.lower()=='n'):
        quit()
